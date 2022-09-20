import openpyxl
import pymysql
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import Font
import pandas as pd

#creating DB connection
class Db_conn:
    BODY_OK = "Plese see attahced file for expiring licenses.\nto stop this service please close 'expiring_licenses' from services.msc"
    BODY_NOT_OK = '''File is opened therefore latest version attached,\nmake sure the file is closed ('lic1.xlsx')
    \nto stop this service please close 'expiring_licenses' from services.msc'''
    def __init__(self):
        # connect to DB
        try:
            self.conn = pymysql.connections.Connection(host='localhost',
                                                     database='your DB name',
                                                     user='root',
                                                     password='enter root pass')
        except Exception:
            print("Connection did not succeed")
        else:
            print("Connection succeeded")
        self.cur = self.conn.cursor()

        # connect to excel file
        try:
            self.wb = openpyxl.load_workbook('lic.xlsx', data_only=True)
        except Exception:
            print('There is no file "lic.xlsx" ')
        else:
            print("Connected to Excel")
        self.ws = self.wb.active
        self.wb_new = openpyxl.Workbook()
        self.ws_new = self.wb_new.active

    def connection(self):
        sqlcopy = '''DROP TABLE IF EXISTS shadowcopy;'''
        self.cur.execute(sqlcopy)
        try:
            self.cur.execute('CREATE TABLE shadowcopy AS SELECT * FROM licenses')
        except Exception:
            print('Could not create shadowcopy')

        sqlreset = '''delete from licenses'''
        self.cur.execute(sqlreset)
        self.cur.execute('alter table licenses auto_increment=1')
        self.conn.commit()
        self.data()

#extract data from excel to DB
    def data(self):
        rows = self.ws.iter_rows(min_row=2, max_row=len(self.ws['A']), min_col=1, max_col=12)
        for row in rows:
            each_row = []
            for cell in row:
                each_row.append(cell.value)
            # print(each_row)
            try:
                company = each_row[0]
                lic_name = each_row[1]
                supplier = each_row[2]
                frequency = each_row[3]
                last_update = each_row[4]
                next_update = each_row[5]
                time_left_delta = each_row[5] - datetime.today()
                time_left = time_left_delta.days +1
                file_path = each_row[7]

                sql = "insert into licenses (company, lic_name, supplier, frequency, last_update,next_update,time_left, file_path) values ('{}', '{}', '{}', '{}', '{}','{}','{}','{}') ".format(company,lic_name,supplier,frequency,last_update,next_update,time_left, file_path)
                self.cur.execute(sql)
                self.conn.commit()
            except Exception:
                pass
#sending warning mail for licenses that about to be expired
    def send_mail(self,body):
        sent_from = "your mail"
        to_mail = 'recipient@gmail.com'
        msg = MIMEMultipart()
        msg['From'] = sent_from
        msg['To'] = to_mail
        msg['Subject'] = 'Expiring licenses'
        # body = 'See attached file for expiring licenses'
        # attach the body with the msg instance
        msg.attach(MIMEText(body, 'plain'))
        filename = ("enter attachment file path")
        attachment = open(filename, "rb")
        # instance of MIMEBase and named as p
        p = MIMEBase('application', 'octet-stream')
        # To change the payload into encoded form
        p.set_payload((attachment).read())
        # encode into base64
        encoders.encode_base64(p)
        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        # attach the instance 'p' to instance 'msg'
        msg.attach(p)
        # creates SMTP session
        s = smtplib.SMTP('smtp.gmail.com', 587)

        # start TLS for security
        s.starttls()

        # Authentication from sql hashed
        s.login(sent_from, "mail pass")

        # Converts the Multipart msg into a string
        text = msg.as_string()

        # sending the mail
        s.sendmail(sent_from, to_mail, text)
        print('Mail sent successfully')
        # terminating the session
        s.quit()
        
#extract data from db to insert excel file
    def get_data_fromDB(self):
        # expiring = ('select * from licenses where time_left <30 order by time_left asc')
        expiring = ('select company,lic_name,supplier,frequency,last_update,next_update,time_left from lics.licenses where time_left <30 order by time_left asc')
        self.cur.execute(expiring)
        self.all_data = self.cur.fetchall()
        return self.all_data

    def insert_to_excel(self):
        self.ws_new.title = "Expiring Licenses"
        font = Font(bold =True)
        self.ws_new['A1'].value = "Company"
        self.ws_new['A1'].font = font
        self.ws_new['B1'].value = "Lic name"
        self.ws_new['B1'].font = font
        self.ws_new['C1'].value = "Supplier"
        self.ws_new['C1'].font = font
        self.ws_new['D1'].value = "Frequency"
        self.ws_new['D1'].font = font
        self.ws_new['E1'].value = "Last Update"
        self.ws_new['E1'].font = font
        self.ws_new['F1'].value = "Next Update"
        self.ws_new['F1'].font = font
        self.ws_new['G1'].value = "Time Left"
        self.ws_new['G1'].font = font

        for row in self.all_data:
            self.ws_new.append(row)
            # company = row[1]
            # lic_name = row[2]
            # supplier = row[3]
            # frequency = row[4]
            # last_update = row[5]
            # next_update = row[6]
            # time_left = row[7]
            # # file_path =str(row[8])
        try:
            self.wb_new.save('lic1.xlsx')
            body = self.BODY_OK
        except PermissionError:
            print("file is opened therefore cannot be saved")
            body = self.BODY_NOT_OK
            self.send_mail(body)
        else:
            self.send_mail(body)
#short version to the purpose of the code without using DB
    def get_excel_with_pandas(self):
        data = pd.read_excel(open('lic.xlsx', 'rb'),
                      sheet_name='lic')
        df = pd.DataFrame(data)
        df1 = df[(df['update_in']<=30)]
        df1.sort_values(by=['update_in'],ascending=True)
        try:
            df1.to_excel("lic1.xlsx", sheet_name="Expiring Licenses")
        except PermissionError:
            print("File is opened, attached latest version")
            body = self.BODY_NOT_OK
            self.send_mail(body)
        else:
            body = self.BODY_OK
            self.send_mail(body)

#driver code
def main():
    db = Db_conn()
    # db.connection()
    # db.get_data_fromDB()
    # db.insert_to_excel()
    db.get_excel_with_pandas()

if __name__=="__main__":
    main()
